
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from scipy.stats import skew, shapiro
from scipy.signal import find_peaks
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from docx import Document
from docx.shared import Inches
from docx.enum.table import WD_TABLE_ALIGNMENT
from docx.enum.text import WD_ALIGN_PARAGRAPH
import io
import os

st.set_page_config(page_title="Análisis Bioestadístico de Peces", layout="wide")
st.title("📊 Análisis Bioestadístico de Población de Peces (Enero-Junio)")

uploaded_file = st.file_uploader("📥 Sube tu archivo Excel de muestreo (.xlsx)", type="xlsx")

if uploaded_file is not None:
    df_excel = pd.read_excel(uploaded_file, sheet_name=0)
    df_excel.to_csv("105.csv", index=False)
    df = df_excel.copy()

    df["Fecha muestreo"] = pd.to_datetime(df["Fecha muestreo"])
    df["Mes"] = df["Fecha muestreo"].dt.to_period("M").astype(str)
    df = df[df["Fecha muestreo"].dt.month <= 6]

    def coef_var(x):
        return x.std() / x.mean() if x.mean() != 0 else 0

    variables = ["Peso", "Longitud", "K"]
    estadisticas = []

    for var in variables:
        resumen = df.groupby("Mes")[var].agg([
            ("Promedio " + var, "mean"),
            ("Coef. Variación " + var, coef_var),
            ("1er Cuartil " + var, lambda x: x.quantile(0.25)),
            ("2º Cuartil (Mediana) " + var, "median"),
            ("3er Cuartil " + var, lambda x: x.quantile(0.75)),
            ("4º Cuartil (Máximo) " + var, "max")
        ])
        estadisticas.append(resumen)

    tabla_final = pd.concat(estadisticas, axis=1).reset_index()
    tabla_final = tabla_final.round(2)

    salida_excel = "estadisticas_mensuales.xlsx"
    tabla_final.to_excel(salida_excel, index=False)

    wb = load_workbook(salida_excel)
    ws = wb.active
    ws.title = "Estadísticas Mensuales"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    for i, col in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(max_length + 2, 30))

    ws.freeze_panes = "A2"
    wb.save(salida_excel)
    st.success("✅ Tabla de estadísticas exportada como Excel")

    st.subheader("📈 Violinplots")
    for var in variables:
        fig = plt.figure(figsize=(10, 4))
        sns.violinplot(x="Mes", y=var, data=df, palette="pastel")
        plt.title(f"Violinplot mensual de {var}")
        st.pyplot(fig)

    st.subheader("📦 Boxplot de Peso con Coeficiente de Variación")
    cv_peso = df.groupby("Mes")["Peso"].apply(coef_var)
    fig = plt.figure(figsize=(10, 4))
    ax = sns.boxplot(x="Mes", y="Peso", data=df, palette="pastel")
    for i, mes in enumerate(sorted(df["Mes"].unique())):
        cv_val = cv_peso.loc[mes]
        plt.text(i, df["Peso"].max() * 1.02, f"CV={cv_val:.2f}", ha='center', fontsize=10, fontweight='bold')
    st.pyplot(fig)

    st.subheader("📊 Histogramas de Peso con Análisis de Sesgo")
    peso_min, peso_max = df["Peso"].min(), df["Peso"].max()
    bin_width = 50
    bins = np.arange(peso_min, peso_max + bin_width, bin_width)
    for mes in sorted(df["Mes"].unique()):
        data_mes = df[df["Mes"] == mes]["Peso"]
        sesgo = skew(data_mes)
        if sesgo > 0.5:
            tipo_sesgo = "Cola moderada a la derecha"
        elif sesgo > 0.1:
            tipo_sesgo = "Cola ligera a la derecha"
        elif sesgo < -0.5:
            tipo_sesgo = "Cola moderada a la izquierda"
        elif sesgo < -0.1:
            tipo_sesgo = "Cola ligera a la izquierda"
        else:
            tipo_sesgo = "Distribución simétrica"
        stat, p_value = shapiro(data_mes)
        normalidad = "Aprox. normal" if p_value > 0.05 else "No normal"
        kde = sns.kdeplot(data_mes, bw_adjust=1)
        x_kde, y_kde = kde.get_lines()[0].get_data()
        peaks, _ = find_peaks(y_kde, height=max(y_kde) * 0.1)
        forma = ["Unimodal", "Bimodal"][len(peaks)-1] if len(peaks) in [1,2] else f"{len(peaks)}-modal"
        plt.clf()
        fig, ax = plt.subplots(figsize=(8, 4))
        sns.histplot(data_mes, bins=bins, kde=True, ax=ax, color="skyblue", edgecolor="black")
        promedio = data_mes.mean()
        ax.axvline(promedio, color="red", linestyle="--", linewidth=2)
        ax.set_title(f"{mes} | {tipo_sesgo} | {forma} | {normalidad} (skew={sesgo:.2f}, p={p_value:.3f})")
        st.pyplot(fig)

    with open("estadisticas_mensuales.xlsx", "rb") as f:
        st.download_button("📥 Descargar tabla Excel", f, file_name="estadisticas_mensuales.xlsx")
